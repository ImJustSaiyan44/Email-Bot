#For Exel Files
from selenium import webdriver
import time
from time import sleep
from Login import username,password,usernameG,passwordG
from selenium.webdriver.common.keys import Keys
import openpyxl
from openpyxl import *
from openpyxl.styles import *

#Exel Style Sheets
fonth = Font(name='Calibri',
size=18,
color='FF000000')
fontI = Font(name='Calibri',
size=18,
color='FF000000')
fill = PatternFill(fill_type='solid',
start_color='9BC2E6')

ExelPath = ('E:\Cnz Student Roster\ExelsheetName.xlsx')
sheet = openpyxl.load_workbook(ExelPath)
ws = sheet.active
Path = executable_path=r"C:\Program Files (x86)\chromedriver.exe"

driver = webdriver.Chrome(Path)

driver.get('https://login.live.com/login.srf?wa=wsignin1.0&rpsnv=13&ct=1615416443&rver=7.0.6737.0&wp=MBI_SSL&wreply=https%3a%2f%2foutlook.live.com%2fowa%2f%3fnlp%3d1%26RpsCsrfState%3d241e3587-f949-05c3-571e-a507dea6d2fe&id=292841&aadredir=1&CBCXT=out&lw=1&fl=dob%2cflname%2cwld')

#Login into outlook
search = driver.find_element_by_xpath("/html/body/div/form[1]/div/div/div[2]/div/div/div[1]/div[2]/div[2]/div/div/div/div[2]/div[2]/div/input[1]")
search.send_keys(username)
search.send_keys(Keys.RETURN)
time.sleep(2)
search = driver.find_element_by_xpath("/html/body/div/form[1]/div/div/div[2]/div/div/div[2]/div[2]/div[2]/div/div[2]/div/div[2]/div/div[2]/input")
search.send_keys(password)
search.send_keys(Keys.RETURN)
time.sleep(2)

#Find continue Button
search = driver.find_element_by_xpath("/html/body/div/form/div/div/div[2]/div[2]/div/div[2]/div/div[3]/div[2]/div/div/div[2]/input")
search.send_keys(Keys.RETURN)
time.sleep(2)

#Open Folder EventBrite 
link = driver.find_element_by_xpath("/html/body/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div[2]/div/div[1]/div/div[2]/div[10]/div")
link.click()
time.sleep(1)

#This is to search through emails
eventbrite = driver.find_elements_by_class_name("_1xP-XmXM1GGHpRKCCeOKjP")#relative class
#after first one get Number from sheet
info = []

#Names
Namelist = []
#Proccesed names
PNames = []
counter = 3

#sort through emails
for eventbriteE in eventbrite:
    A = 'A' + str(counter)
    B = 'B' + str(counter)
    C = 'C' + str(counter)
    D = 'D' + str(counter)
    E = 'E' + str(counter)
    F = 'F' + str(counter)
    H = 'H' + str(counter)
    I = 'I' + str(counter)
    J = 'J' + str(counter)
    K = 'K' + str(counter)
    #Select email
    eventbriteE.click()
    name = eventbriteE.find_element_by_xpath('//*[@id="ReadingPaneContainerId"]/div/div/div/div[2]/div/div[1]/div/div/div/div[3]/div/div/div/div[3]/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr[3]/td/table/tbody/tr/td[2]/table/tbody/tr[6]/td/span').text
    Namelist.append(name)
    classt = eventbriteE.find_element_by_xpath('//*[@id="ReadingPaneContainerId"]/div/div/div/div[2]/div/div[1]/div/div/div/div[3]/div/div/div/div[3]/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr[3]/td/table/tbody/tr/td[2]/table/tbody/tr[3]/td/span/a').text
    email = eventbriteE.find_element_by_xpath('//*[@id="ReadingPaneContainerId"]/div/div/div/div[2]/div/div[1]/div/div/div/div[3]/div/div/div/div[3]/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr[3]/td/table/tbody/tr/td[2]/table/tbody/tr[7]/td/span/a').text
    info.append(name)
    info.append(classt)
    info.append(email)
    #TO_DO Get link Of order number
    orderNum = eventbriteE.find_element_by_xpath('//*[@id="ReadingPaneContainerId"]/div/div/div/div[2]/div/div[1]/div/div/div/div[3]/div/div/div/div[3]/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr[3]/td/table/tbody/tr/td[2]/table/tbody/tr[8]/td/span/a').text
    info.append(orderNum)
   
    time.sleep(2)
    for name in Namelist:
        if (name in PNames):
            print('Duplicate Name')
            #TO-DO Add other classes
            
        else:
            PNames.append(name)
            counter += 1
            ws[A]= info[0]
            ws[B]= info[1]
            ws[C]= info[2]
            ws[J]= info[3]
            
        
    print(counter)       
    
    #Cell count Value
    ws['B1'] = counter
    # Style For Counter
    c = ws['B1']
    c.font = fonth
    c.fill = fill
    sheet.save(ExelPath)
   
    
    info = []
    
